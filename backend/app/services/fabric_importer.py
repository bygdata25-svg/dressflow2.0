from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from fastapi import HTTPException, UploadFile
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.fabric import Fabric
from app.models.fabric_roll import FabricRoll
from app.models.supplier import Supplier


EXPECTED_COLUMNS = {
    "CODIGO",
    "NOMBRE TELA PROV Y COD",
    "TELA BASE",
    "COD TELA BASE",
    "COLOR GOROF",
    "COLOR PROV",
    "PROV",
    "COD PROV",
    "TOTAL",
    "CANTIDAD CORTE 1",
    "CANT CORTE 2",
    "CANT CORT 3",
    "CANT ROLLO 1",
    "CANT ROLLO 2",
    "CANT ROLLO 3",
    "CANT ROLLO 4",
    "RETAZOS",
    "PRECIO USS",
    "UBICACION",
    "COMPOSICION",
    "ANCHO METROS",
    "PESO EN GRAMOS",
    "RINDE KILOS",
    "ORIGEN",
}


CUT_COLUMNS = [
    ("CANTIDAD CORTE 1", "CUT", "CORTE_1"),
    ("CANT CORTE 2", "CUT", "CORTE_2"),
    ("CANT CORT 3", "CUT", "CORTE_3"),
]

ROLL_COLUMNS = [
    ("CANT ROLLO 1", "ROLL", "ROLLO_1"),
    ("CANT ROLLO 2", "ROLL", "ROLLO_2"),
    ("CANT ROLLO 3", "ROLL", "ROLLO_3"),
    ("CANT ROLLO 4", "ROLL", "ROLLO_4"),
]


@dataclass
class ImportStats:
    total_rows: int = 0
    created_suppliers: int = 0
    reused_suppliers: int = 0
    created_fabrics: int = 0
    reused_fabrics: int = 0
    created_rolls: int = 0
    skipped_rolls: int = 0
    errors: list[dict[str, Any]] | None = None

    def __post_init__(self) -> None:
        if self.errors is None:
            self.errors = []


def _clean_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null"}:
        return None
    return text


def _clean_key(value: Any) -> str:
    text = _clean_str(value) or ""
    return re.sub(r"\s+", " ", text).strip().upper()


def _to_decimal(value: Any) -> Decimal | None:
    text = _clean_str(value)
    if text is None:
        return None

    text = text.replace("$", "").replace("U$S", "").replace("USD", "").strip()

    # soporta "1.234,56" y "1234.56"
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    else:
        text = text.replace(",", ".")

    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _to_bool(value: Any) -> bool:
    text = (_clean_str(value) or "").strip().lower()
    return text in {"si", "sí", "yes", "y", "true", "1", "x"}


def _slug_piece_code(base_code: str, suffix: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9_-]+", "-", base_code.strip())
    safe = re.sub(r"-{2,}", "-", safe).strip("-")
    return f"{safe}-{suffix}"[:100]


def _normalize_row_keys(row: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for k, v in row.items():
        normalized[_clean_key(k)] = v
    return normalized


async def _read_rows_from_upload(file: UploadFile) -> list[dict[str, Any]]:
    filename = file.filename or "archivo"
    ext = Path(filename).suffix.lower()

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="El archivo está vacío.")

    if ext == ".csv":
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        return [_normalize_row_keys(row) for row in reader]

    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(
                status_code=500,
                detail="Falta openpyxl para procesar archivos .xlsx.",
            ) from exc

        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [_clean_key(h) for h in rows[0]]
        result: list[dict[str, Any]] = []
        for row in rows[1:]:
            row_dict = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
            result.append(row_dict)
        return result

    if ext == ".ods":
        try:
            import pandas as pd
        except ImportError as exc:
            raise HTTPException(
                status_code=400,
                detail="Para importar .ods instalá pandas y odfpy, o exportá la planilla a .xlsx.",
            ) from exc

        try:
            df = pd.read_excel(io.BytesIO(content), engine="odf")
        except Exception as exc:
            raise HTTPException(
                status_code=400,
                detail=f"No pude leer el archivo .ods: {exc}",
            ) from exc

        return [_normalize_row_keys(row) for row in df.to_dict(orient="records")]

    raise HTTPException(
        status_code=400,
        detail="Formato no soportado. Usá .xlsx, .csv o .ods.",
    )


def _validate_columns(rows: list[dict[str, Any]]) -> None:
    if not rows:
        raise HTTPException(status_code=400, detail="La planilla no tiene filas para importar.")

    columns = set(rows[0].keys())
    missing = sorted(EXPECTED_COLUMNS - columns)
    if missing:
        raise HTTPException(
            status_code=400,
            detail={
                "message": "Faltan columnas esperadas en la planilla.",
                "missing_columns": missing,
            },
        )


def _find_supplier(db: Session, tenant_id, supplier_code: str | None, supplier_name: str | None) -> Supplier | None:
    if supplier_code:
        stmt = select(Supplier).where(
            Supplier.tenant_id == tenant_id,
            Supplier.supplier_code == supplier_code,
            Supplier.deleted_at.is_(None),
        )
        supplier = db.execute(stmt).scalar_one_or_none()
        if supplier:
            return supplier

    if supplier_name:
        stmt = select(Supplier).where(
            Supplier.tenant_id == tenant_id,
            Supplier.name == supplier_name,
            Supplier.deleted_at.is_(None),
        )
        return db.execute(stmt).scalar_one_or_none()

    return None


def _find_fabric(db: Session, tenant_id, code: str | None, name: str | None, color: str | None) -> Fabric | None:
    if code:
        stmt = select(Fabric).where(
            Fabric.tenant_id == tenant_id,
            Fabric.code == code,
            Fabric.deleted_at.is_(None),
        )
        fabric = db.execute(stmt).scalar_one_or_none()
        if fabric:
            return fabric

    if name:
        stmt = select(Fabric).where(
            Fabric.tenant_id == tenant_id,
            Fabric.name == name,
            Fabric.color == color,
            Fabric.deleted_at.is_(None),
        )
        return db.execute(stmt).scalar_one_or_none()

    return None


def _roll_exists(db: Session, tenant_id, roll_code: str) -> bool:
    stmt = select(FabricRoll.id).where(
        FabricRoll.tenant_id == tenant_id,
        FabricRoll.roll_code == roll_code,
        FabricRoll.deleted_at.is_(None),
    )
    return db.execute(stmt).scalar_one_or_none() is not None


def _build_piece_rows(row: dict[str, Any], fabric_code: str, price_per_meter: Decimal | None, location: str | None):
    pieces: list[dict[str, Any]] = []

    for col_name, piece_type, legacy_slot in CUT_COLUMNS + ROLL_COLUMNS:
        meters = _to_decimal(row.get(col_name))
        if meters is None or meters <= 0:
            continue

        suffix = legacy_slot.replace("_", "")
        roll_code = _slug_piece_code(fabric_code, suffix)

        pieces.append(
            {
                "roll_code": roll_code,
                "piece_type": piece_type,
                "legacy_slot": legacy_slot,
                "initial_length": meters,
                "current_length": meters,
                "reserved_length": Decimal("0"),
                "price_per_meter": price_per_meter,
                "location": location,
                "is_scrap": False,
                "status": "AVAILABLE",
            }
        )

    return pieces


async def import_fabrics_file(
    db: Session,
    tenant_id,
    file: UploadFile,
    *,
    dry_run: bool = True,
    import_batch: str | None = None,
) -> dict[str, Any]:
    rows = await _read_rows_from_upload(file)
    _validate_columns(rows)

    stats = ImportStats(total_rows=len(rows))
    preview_rows: list[dict[str, Any]] = []

    for index, raw_row in enumerate(rows, start=2):
        try:
            codigo = _clean_str(raw_row.get("CODIGO"))
            nombre_tela = _clean_str(raw_row.get("NOMBRE TELA PROV Y COD")) or codigo or f"TELA FILA {index}"
            tela_base = _clean_str(raw_row.get("TELA BASE"))
            cod_tela_base = _clean_str(raw_row.get("COD TELA BASE"))
            color_gorof = _clean_str(raw_row.get("COLOR GOROF"))
            color_prov = _clean_str(raw_row.get("COLOR PROV"))
            proveedor_nombre = _clean_str(raw_row.get("PROV"))
            proveedor_codigo = _clean_str(raw_row.get("COD PROV"))
            ubicacion = _clean_str(raw_row.get("UBICACION"))
            composicion = _clean_str(raw_row.get("COMPOSICION"))
            origen = _clean_str(raw_row.get("ORIGEN"))
            precio_uss = _to_decimal(raw_row.get("PRECIO USS"))
            ancho_metros = _to_decimal(raw_row.get("ANCHO METROS"))
            peso_gramos = _to_decimal(raw_row.get("PESO EN GRAMOS"))
            rinde_kilos = _to_decimal(raw_row.get("RINDE KILOS"))
            has_scraps = _to_bool(raw_row.get("RETAZOS"))

            supplier = _find_supplier(db, tenant_id, proveedor_codigo, proveedor_nombre)
            supplier_created = False
            if supplier is None and proveedor_nombre:
                supplier = Supplier(
                    tenant_id=tenant_id,
                    name=proveedor_nombre,
                    supplier_code=proveedor_codigo,
                    origin=origen,
                    supplier_type="FABRIC_SUPPLIER",
                    is_active=True,
                )
                db.add(supplier)
                db.flush()
                supplier_created = True
                stats.created_suppliers += 1
            elif supplier is not None:
                stats.reused_suppliers += 1

            fabric = _find_fabric(db, tenant_id, codigo, nombre_tela, color_gorof)
            fabric_created = False
            if fabric is None:
                fabric = Fabric(
                    tenant_id=tenant_id,
                    supplier_id=supplier.id if supplier else None,
                    name=nombre_tela,
                    fabric_type=tela_base,
                    color=color_gorof,
                    notes=None,
                    photo_url=None,
                    code=codigo,
                    base_name=tela_base,
                    base_code=cod_tela_base,
                    supplier_color=color_prov,
                    supplier_reference=nombre_tela,
                    composition=composicion,
                    origin=origen,
                    width_meters=ancho_metros,
                    weight_grams=peso_gramos,
                    yield_kilos=rinde_kilos,
                    default_location=ubicacion,
                    has_scraps=has_scraps,
                    is_active=True,
                )
                db.add(fabric)
                db.flush()
                fabric_created = True
                stats.created_fabrics += 1
            else:
                stats.reused_fabrics += 1

            base_code_for_rolls = codigo or fabric.code or f"FABRIC-{index}"
            piece_rows = _build_piece_rows(
                raw_row,
                fabric_code=base_code_for_rolls,
                price_per_meter=precio_uss,
                location=ubicacion,
            )

            preview_rows.append(
                {
                    "row_number": index,
                    "codigo": codigo,
                    "fabric_name": nombre_tela,
                    "supplier_name": proveedor_nombre,
                    "supplier_created": supplier_created,
                    "fabric_created": fabric_created,
                    "pieces_detected": len(piece_rows),
                    "has_scraps": has_scraps,
                }
            )

            for piece in piece_rows:
                if _roll_exists(db, tenant_id, piece["roll_code"]):
                    stats.skipped_rolls += 1
                    continue

                roll = FabricRoll(
                    tenant_id=tenant_id,
                    fabric_id=fabric.id,
                    supplier_id=supplier.id if supplier else None,
                    roll_code=piece["roll_code"],
                    piece_type=piece["piece_type"],
                    legacy_slot=piece["legacy_slot"],
                    initial_length=piece["initial_length"],
                    current_length=piece["current_length"],
                    reserved_length=piece["reserved_length"],
                    unit="meters",
                    status=piece["status"],
                    price_per_meter=piece["price_per_meter"],
                    currency="USD" if precio_uss is not None else None,
                    purchase_date=None,
                    location=piece["location"],
                    is_scrap=piece["is_scrap"],
                    is_active=True,
                    import_batch=import_batch,
                    import_row_number=index,
                    notes=None,
                )
                db.add(roll)
                stats.created_rolls += 1

        except Exception as exc:
            stats.errors.append(
                {
                    "row_number": index,
                    "codigo": raw_row.get("CODIGO"),
                    "error": str(exc),
                }
            )

    if dry_run:
        db.rollback()
    else:
        if stats.errors:
            db.rollback()
            raise HTTPException(
                status_code=400,
                detail={
                    "message": "La importación encontró errores y fue cancelada.",
                    "errors": stats.errors,
                },
            )
        db.commit()

    return {
        "dry_run": dry_run,
        "import_batch": import_batch,
        "stats": {
            "total_rows": stats.total_rows,
            "created_suppliers": stats.created_suppliers,
            "reused_suppliers": stats.reused_suppliers,
            "created_fabrics": stats.created_fabrics,
            "reused_fabrics": stats.reused_fabrics,
            "created_rolls": stats.created_rolls,
            "skipped_rolls": stats.skipped_rolls,
            "errors_count": len(stats.errors),
        },
        "errors": stats.errors,
        "preview": preview_rows[:50],
    }
