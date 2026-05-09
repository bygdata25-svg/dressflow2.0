from io import BytesIO
from decimal import Decimal
from datetime import date, datetime
import pandas as pd
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import text, select, or_, desc
from sqlalchemy.orm import Session
from app.api.deps import get_current_membership

from app.api.deps import require_roles
from app.core.database import get_db
from app.models.dress_sale import DressSale
from app.models.dress import Dress
from app.models.customer import Customer
from app.models.accessory_sale import AccessorySale
from app.models.accessory import Accessory

router = APIRouter(prefix="/reports", tags=["reports"])


# =========================================================
# HELPERS
# =========================================================

def _excel_response(workbook: Workbook, filename: str) -> StreamingResponse:
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _apply_header_style(ws, headers: list[str]) -> None:
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="3D3648")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, _header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")



# =========================================================
# I18N HELPERS FOR EXCEL EXPORTS
# =========================================================

REPORTS_I18N = {
    "es": {
        "stock_valuation": {
            "sheet": "Stock valorizado",
            "filename": "stock_valorizado.xlsx",
            "headers": [
                "Tela", "Color", "Stock disponible (mts)", "Precio promedio x metro", "Valor total",
            ],
            "total": "TOTAL",
        },
        "fabric_movements": {
            "sheet": "Movimientos tela",
            "filename": "movimientos_tela.xlsx",
            "headers": [
                "Fecha", "Tipo", "Tela", "Color", "Rollo", "Cantidad", "Referencia", "Motivo", "Notas",
            ],
            "types": {"IN": "Entrada", "OUT": "Salida", "ADJUST": "Ajuste", "ADJUSTMENT": "Ajuste"},
            "reasons": {
                "PRODUCTION_ISSUE": "Orden de producción",
                "PRODUCTION_RETURN": "Devolución de producción",
            },
        },
        "loans": {
            "sheet": "Préstamos y alquileres",
            "filename": "prestamos_alquileres.xlsx",
            "headers": [
                "Tipo", "Fecha inicio", "Fecha vencimiento", "Fecha devolución", "Estado",
                "Código vestido", "Nombre vestido", "Cliente", "Valor", "Notas",
            ],
            "types": {"LOAN": "Préstamo", "RENTAL": "Alquiler"},
            "status": {"ACTIVE": "Activo", "LATE": "Vencido", "RETURNED": "Devuelto"},
        },
        "production_costs": {
            "sheet": "Costos producción",
            "filename": "costos_produccion.xlsx",
            "headers": [
                "Orden", "Vestido", "Taller", "Estado", "Prioridad", "Entrega", "Planificado", "Producido",
                "Materiales est.", "Materiales real", "Mano de obra", "Adicional", "Total est.",
                "Total real", "Unitario est.", "Unitario real", "Moneda",
            ],
            "status": {
                "DRAFT": "Borrador", "PLANNED": "Planificada", "IN_PROGRESS": "En proceso",
                "DONE": "Finalizada", "CANCELLED": "Cancelada", "COMPLETED": "Completada",
            },
            "priority": {"LOW": "Baja", "NORMAL": "Normal", "MEDIUM": "Media", "HIGH": "Alta", "URGENT": "Urgente"},
        },
        "sales": {
            "sheet": "Ventas",
            "filename": "ventas.xlsx",
            "headers": [
                "N° Venta", "Fecha", "Código vestido", "Vestido", "Cliente", "Precio", "Moneda",
                "Método de pago", "Estado", "Notas",
            ],
            "total": "TOTAL",
        },
        "accessory_sales": {
            "sheet": "Ventas accesorios",
            "filename": "ventas_accesorios.xlsx",
            "headers": [
                "N° Venta", "Fecha", "Código accesorio", "Accesorio", "Cliente", "Cantidad",
                "Precio unitario", "Total", "Moneda", "Método de pago", "Estado", "Notas",
            ],
            "total": "TOTAL",
        },
        "sales_unified": {
            "sheet": "Ventas unificadas",
            "filename": "ventas_unificadas.xlsx",
            "headers": [
                "N° Venta", "Fecha", "Cliente", "Moneda cabecera", "Estado", "Total ítems ARS",
                "Total ítems USD", "Pagado ARS", "Pagado USD", "Ítems", "Pagos", "Notas",
            ],
            "total": "TOTALES",
            "item_types": {"DRESS": "Vestido", "ACCESSORY": "Accesorio", "TRIM": "Avío"},
            "item_fallback": "Ítem",
        },
        "sale_status": {"COMPLETED": "Completada", "CANCELLED": "Cancelada", "DRAFT": "Borrador", "PAID": "Pagada", "PARTIAL": "Parcial", "PENDING": "Pendiente"},
        "payment_methods": {"CASH": "EFECTIVO", "TRANSFER": "TRANSFERENCIA", "CARD": "TARJETA DE CRÉDITO", "MERCADOPAGO": "MERCADO PAGO"},
    },
    "en": {
        "stock_valuation": {
            "sheet": "Stock valuation",
            "filename": "stock_valuation.xlsx",
            "headers": [
                "Fabric", "Color", "Available stock (m)", "Average price per meter", "Total value",
            ],
            "total": "TOTAL",
        },
        "fabric_movements": {
            "sheet": "Fabric movements",
            "filename": "fabric_movements.xlsx",
            "headers": [
                "Date", "Type", "Fabric", "Color", "Roll", "Quantity", "Reference", "Reason", "Notes",
            ],
            "types": {"IN": "In", "OUT": "Out", "ADJUST": "Adjustment", "ADJUSTMENT": "Adjustment"},
            "reasons": {
                "PRODUCTION_ISSUE": "Production order",
                "PRODUCTION_RETURN": "Production return",
            },
        },
        "loans": {
            "sheet": "Loans and rentals",
            "filename": "loans_rentals.xlsx",
            "headers": [
                "Type", "Start date", "Due date", "Return date", "Status",
                "Dress code", "Dress name", "Customer", "Amount", "Notes",
            ],
            "types": {"LOAN": "Loan", "RENTAL": "Rental"},
            "status": {"ACTIVE": "Active", "LATE": "Late", "RETURNED": "Returned"},
        },
        "production_costs": {
            "sheet": "Production costs",
            "filename": "production_costs.xlsx",
            "headers": [
                "Order", "Dress", "Workshop", "Status", "Priority", "Due date", "Planned", "Produced",
                "Est. materials", "Actual materials", "Labor", "Additional", "Est. total",
                "Actual total", "Est. unit", "Actual unit", "Currency",
            ],
            "status": {
                "DRAFT": "Draft", "PLANNED": "Planned", "IN_PROGRESS": "In progress",
                "DONE": "Done", "CANCELLED": "Cancelled", "COMPLETED": "Completed",
            },
            "priority": {"LOW": "Low", "NORMAL": "Normal", "MEDIUM": "Medium", "HIGH": "High", "URGENT": "Urgent"},
        },
        "sales": {
            "sheet": "Sales",
            "filename": "sales.xlsx",
            "headers": [
                "Sale #", "Date", "Dress code", "Dress", "Customer", "Price", "Currency",
                "Payment method", "Status", "Notes",
            ],
            "total": "TOTAL",
        },
        "accessory_sales": {
            "sheet": "Accessory sales",
            "filename": "accessory_sales.xlsx",
            "headers": [
                "Sale #", "Date", "Accessory code", "Accessory", "Customer", "Quantity",
                "Unit price", "Total", "Currency", "Payment method", "Status", "Notes",
            ],
            "total": "TOTAL",
        },
        "sales_unified": {
            "sheet": "Unified sales",
            "filename": "unified_sales.xlsx",
            "headers": [
                "Sale #", "Date", "Customer", "Header currency", "Status", "Items total ARS",
                "Items total USD", "Paid ARS", "Paid USD", "Items", "Payments", "Notes",
            ],
            "total": "TOTALS",
            "item_types": {"DRESS": "Dress", "ACCESSORY": "Accessory", "TRIM": "Trim"},
            "item_fallback": "Item",
        },
        "sale_status": {"COMPLETED": "Completed", "CANCELLED": "Cancelled", "DRAFT": "Draft", "PAID": "Paid", "PARTIAL": "Partial", "PENDING": "Pending"},
        "payment_methods": {"CASH": "CASH", "TRANSFER": "TRANSFER", "CARD": "CREDIT CARD", "MERCADOPAGO": "MERCADO PAGO"},
    },
}


def _lang(lang: str | None) -> str:
    return "en" if str(lang or "es").lower().startswith("en") else "es"


def _tx(section: str, lang: str | None) -> dict:
    return REPORTS_I18N[_lang(lang)][section]


def _label(section: str, group: str, value: str | None, lang: str | None, fallback: str = "") -> str:
    raw = (value or "").upper().strip()
    if not raw:
        return fallback
    return REPORTS_I18N[_lang(lang)].get(section, {}).get(group, {}).get(raw, value or fallback)


def _movement_type_label(value: str | None, lang: str | None = "es") -> str:
    return _tx("fabric_movements", lang)["types"].get((value or "").upper(), value or "")


def _loan_status_label(value: str | None, lang: str | None = "es") -> str:
    return _tx("loans", lang)["status"].get((value or "").upper(), value or "")

def _loan_status_fill(value: str | None) -> PatternFill:
    raw = (value or "").upper()
    if raw == "ACTIVE":
        return PatternFill("solid", fgColor="ECFDF3")
    if raw == "LATE":
        return PatternFill("solid", fgColor="FDECEC")
    if raw == "RETURNED":
        return PatternFill("solid", fgColor="F4F4F5")
    return PatternFill("solid", fgColor="FFFFFF")


def _loan_status_font(value: str | None) -> Font:
    raw = (value or "").upper()
    if raw == "ACTIVE":
        return Font(color="027A48", bold=True)
    if raw == "LATE":
        return Font(color="B42318", bold=True)
    if raw == "RETURNED":
        return Font(color="52525B", bold=True)
    return Font(color="000000")



def _sale_status_label(value: str | None, lang: str | None = "es") -> str:
    return REPORTS_I18N[_lang(lang)]["sale_status"].get((value or "").upper().strip(), value or "")


def _sale_status_fill(value: str | None) -> PatternFill:
    raw = (value or "").upper()
    if raw in ("COMPLETED", "PAID"):
        return PatternFill("solid", fgColor="ECFDF3")
    if raw == "CANCELLED":
        return PatternFill("solid", fgColor="F4F4F5")
    if raw in ("PARTIAL", "PENDING"):
        return PatternFill("solid", fgColor="FEF3C7")
    return PatternFill("solid", fgColor="FFFFFF")


def _sale_status_font(value: str | None) -> Font:
    raw = (value or "").upper()
    if raw in ("COMPLETED", "PAID"):
        return Font(color="027A48", bold=True)
    if raw == "CANCELLED":
        return Font(color="52525B", bold=True)
    if raw in ("PARTIAL", "PENDING"):
        return Font(color="92400E", bold=True)
    return Font(color="000000")


def _payment_method_label(value: str | None, lang: str | None = "es") -> str:
    raw = (value or "").upper().strip()
    return REPORTS_I18N[_lang(lang)]["payment_methods"].get(raw, value or "")



def _fabric_movement_note(row, lang: str | None = "es") -> str | None:
    """
    Devuelve una nota legible para reportes de movimientos de tela.
    En salidas por producción, prioriza la orden de producción por sobre el texto técnico.
    """
    movement_reason = (row["movement_reason"] or "").upper()
    reference = row["reference"]
    t = _tx("fabric_movements", lang)

    if movement_reason == "PRODUCTION_ISSUE" and reference:
        ref = str(reference).replace("Production Order", "").strip()
        label = t["reasons"].get("PRODUCTION_ISSUE", "Orden de producción")
        return f"{label} {ref}" if ref else label

    if movement_reason == "PRODUCTION_RETURN" and reference:
        ref = str(reference).replace("Production Order", "").strip()
        label = t["reasons"].get("PRODUCTION_RETURN", "Devolución de producción")
        return f"{label} {ref}" if ref else label

    return row["notes"]


# =========================================================
# STOCK VALORIZADO
# =========================================================

def _build_stock_valuation_query(search: str | None = None):
    sql = """
        SELECT
            f.id,
            f.name,
            f.color,
            COALESCE(SUM(fr.current_length - fr.reserved_length), 0) AS total_stock_meters,
            COALESCE(
                SUM(
                    (fr.current_length - fr.reserved_length) * COALESCE(fr.price_per_meter, 0)
                ),
                0
            ) AS total_value
        FROM fabrics f
        LEFT JOIN fabric_rolls fr
            ON fr.fabric_id = f.id
           AND fr.deleted_at IS NULL
           AND fr.status = 'AVAILABLE'
        WHERE f.tenant_id = :tenant_id
          AND f.deleted_at IS NULL
    """

    params: dict[str, object] = {}

    if search:
        sql += """
          AND (
            LOWER(f.name) LIKE :search
            OR LOWER(COALESCE(f.color, '')) LIKE :search
          )
        """
        params["search"] = f"%{search.strip().lower()}%"

    sql += """
        GROUP BY f.id, f.name, f.color
        ORDER BY f.name ASC, f.color ASC
    """

    return text(sql), params


@router.get("/stock-valuation")
def stock_valuation_report(
    db: Session = Depends(get_db),
    membership=Depends(require_roles("admin", "manager", "staff")),
    search: str | None = Query(default=None),
):
    sql, extra_params = _build_stock_valuation_query(search=search)

    rows = db.execute(
        sql,
        {"tenant_id": membership.tenant_id, **extra_params},
    ).mappings().all()

    items = []
    grand_total = Decimal("0")

    for row in rows:
        total_value = Decimal(str(row["total_value"] or 0))
        total_stock_meters = Decimal(str(row["total_stock_meters"] or 0))
        grand_total += total_value

        average_price = (
            float(total_value / total_stock_meters)
            if total_stock_meters > 0
            else 0
        )

        items.append(
            {
                "id": str(row["id"]),
                "name": row["name"],
                "color": row["color"],
                "total_stock_meters": float(total_stock_meters),
                "average_price_per_meter": average_price,
                "total_value": float(total_value),
            }
        )

    return {
        "items": items,
        "total": len(items),
        "grand_total": float(grand_total),
    }


@router.get("/stock-valuation/export")
def export_stock_valuation_report(
    db: Session = Depends(get_db),
    membership=Depends(require_roles("admin", "manager", "staff")),
    search: str | None = Query(default=None),
    lang: str = Query(default="es"),
):
    sql, extra_params = _build_stock_valuation_query(search=search)

    rows = db.execute(
        sql,
        {"tenant_id": membership.tenant_id, **extra_params},
    ).mappings().all()

    tr = _tx("stock_valuation", lang)

    wb = Workbook()
    ws = wb.active
    ws.title = tr["sheet"]

    _apply_header_style(ws, tr["headers"])

    grand_total = Decimal("0")

    for row in rows:
        total_value = Decimal(str(row["total_value"] or 0))
        total_stock_meters = Decimal(str(row["total_stock_meters"] or 0))
        grand_total += total_value

        average_price = (
            float(total_value / total_stock_meters)
            if total_stock_meters > 0
            else 0
        )

        ws.append(
            [
                row["name"],
                row["color"],
                float(total_stock_meters),
                average_price,
                float(total_value),
            ]
        )

    total_row_idx = ws.max_row + 2
    ws.cell(row=total_row_idx, column=4, value=tr["total"])
    ws.cell(row=total_row_idx, column=4).font = Font(bold=True)
    ws.cell(row=total_row_idx, column=4).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=total_row_idx, column=5, value=float(grand_total))
    ws.cell(row=total_row_idx, column=5).font = Font(bold=True)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=5):
        for cell in row:
            cell.number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 18

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{ws.max_row}"

    return _excel_response(wb, tr["filename"])


# =========================================================
# MOVIMIENTOS DE TELA
# =========================================================

def _build_fabric_movements_query(
    search: str | None = None,
    movement_type: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
):
    sql = """
        SELECT
            fm.id,
            fm.created_at,
            fm.type,
            fm.quantity,
            fm.reference,
            fm.notes,
            fm.movement_reason,
            fr.roll_code,
            f.name AS fabric_name,
            f.color AS fabric_color
        FROM fabric_movements fm
        LEFT JOIN fabric_rolls fr
            ON fr.id = fm.fabric_roll_id
        LEFT JOIN fabrics f
            ON f.id = fr.fabric_id
        WHERE fm.tenant_id = :tenant_id
    """

    params: dict[str, object] = {}

    if movement_type:
        sql += " AND fm.type = :movement_type"
        params["movement_type"] = movement_type.strip().upper()

    if date_from:
        sql += " AND DATE(fm.created_at) >= :date_from"
        params["date_from"] = date_from

    if date_to:
        sql += " AND DATE(fm.created_at) <= :date_to"
        params["date_to"] = date_to

    if search:
        sql += """
          AND (
            LOWER(COALESCE(f.name, '')) LIKE :search
            OR LOWER(COALESCE(f.color, '')) LIKE :search
            OR LOWER(COALESCE(fr.roll_code, '')) LIKE :search
            OR LOWER(COALESCE(fm.notes, '')) LIKE :search
            OR LOWER(COALESCE(fm.reference, '')) LIKE :search
            OR LOWER(COALESCE(fm.movement_reason, '')) LIKE :search
          )
        """
        params["search"] = f"%{search.strip().lower()}%"

    sql += """
        ORDER BY fm.created_at DESC
    """

    return text(sql), params


@router.get("/fabric-movements")
def fabric_movements_report(
    db: Session = Depends(get_db),
    membership=Depends(require_roles("admin", "manager", "staff")),
    search: str | None = Query(default=None),
    movement_type: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
):
    sql, extra_params = _build_fabric_movements_query(
        search=search,
        movement_type=movement_type,
        date_from=date_from,
        date_to=date_to,
    )

    rows = db.execute(
        sql,
        {"tenant_id": membership.tenant_id, **extra_params},
    ).mappings().all()

    items = [
        {
            "id": str(row["id"]),
            "created_at": row["created_at"].isoformat() if row["created_at"] else None,
            "type": row["type"],
            "quantity": float(row["quantity"] or 0),
            "reference": row["reference"],
            "notes": _fabric_movement_note(row),
            "movement_reason": row["movement_reason"],
            "fabric_name": row["fabric_name"],
            "fabric_color": row["fabric_color"],
            "roll_code": row["roll_code"],
        }
        for row in rows
    ]

    return {
        "items": items,
        "total": len(items),
    }


@router.get("/fabric-movements/export")
def export_fabric_movements_report(
    db: Session = Depends(get_db),
    membership=Depends(require_roles("admin", "manager", "staff")),
    search: str | None = Query(default=None),
    movement_type: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    lang: str = Query(default="es"),
):
    sql, extra_params = _build_fabric_movements_query(
        search=search,
        movement_type=movement_type,
        date_from=date_from,
        date_to=date_to,
    )

    rows = db.execute(
        sql,
        {"tenant_id": membership.tenant_id, **extra_params},
    ).mappings().all()

    tr = _tx("fabric_movements", lang)

    wb = Workbook()
    ws = wb.active
    ws.title = tr["sheet"]

    _apply_header_style(ws, tr["headers"])

    for row in rows:
        ws.append(
            [
                row["created_at"].strftime("%Y-%m-%d %H:%M") if row["created_at"] else "",
                _movement_type_label(row["type"], lang),
                row["fabric_name"],
                row["fabric_color"],
                row["roll_code"],
                float(row["quantity"] or 0),
                row["reference"],
                row["movement_reason"],
                _fabric_movement_note(row, lang),
            ]
        )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
        for cell in row:
            cell.number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 30

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{ws.max_row}"

    return _excel_response(wb, tr["filename"])


# =========================================================
# PRESTAMOS / ALQUILERES# =========================================================
# PRESTAMOS / ALQUILERES
# =========================================================


def _loan_type_label(value: str | None, lang: str | None = "es") -> str:
    raw = (value or "LOAN").upper()
    return _tx("loans", lang)["types"].get(raw, value or "")


def _build_loans_query(
    search: str | None = None,
    status: str | None = None,
    loan_type: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
):
    sql = """
        SELECT
            l.id,
            l.start_date,
            l.expected_return_date,
            l.actual_return_date,
            l.status,
            l.loan_type,
            l.amount,
            l.notes,
            d.code AS dress_code,
            d.name AS dress_name,
            COALESCE(c.first_name, '') || ' ' || COALESCE(c.last_name, '') AS customer_name
        FROM loans l
        LEFT JOIN dresses d
            ON d.id = l.dress_id
        LEFT JOIN customers c
            ON c.id = l.customer_id
        WHERE l.tenant_id = :tenant_id
          AND l.deleted_at IS NULL
    """

    params: dict[str, object] = {}

    if date_from:
        sql += " AND DATE(l.start_date) >= :date_from"
        params["date_from"] = date_from

    if date_to:
        sql += " AND DATE(l.start_date) <= :date_to"
        params["date_to"] = date_to

    if search:
        sql += """
          AND (
            LOWER(COALESCE(d.code, '')) LIKE :search
            OR LOWER(COALESCE(d.name, '')) LIKE :search
            OR LOWER(COALESCE(c.first_name, '')) LIKE :search
            OR LOWER(COALESCE(c.last_name, '')) LIKE :search
            OR LOWER(COALESCE(l.notes, '')) LIKE :search
          )
        """
        params["search"] = f"%{search.strip().lower()}%"

    if loan_type:
        sql += " AND l.loan_type = :loan_type"
        params["loan_type"] = loan_type.strip().upper()

    if status:
        normalized_status = status.strip().upper()

        if normalized_status == "RETURNED":
            sql += " AND l.actual_return_date IS NOT NULL"
        elif normalized_status == "LATE":
            sql += """
              AND l.actual_return_date IS NULL
              AND l.expected_return_date IS NOT NULL
              AND l.expected_return_date < CURRENT_DATE
            """
        elif normalized_status == "ACTIVE":
            sql += """
              AND l.actual_return_date IS NULL
              AND (
                l.expected_return_date IS NULL
                OR l.expected_return_date >= CURRENT_DATE
              )
            """
        else:
            sql += " AND l.status = :status"
            params["status"] = normalized_status

    sql += """
        ORDER BY l.start_date DESC, l.expected_return_date DESC
    """

    return text(sql), params


@router.get("/loans")
def loans_report(
    db: Session = Depends(get_db),
    membership=Depends(require_roles("admin", "manager", "staff")),
    search: str | None = Query(default=None),
    status: str | None = Query(default=None),
    loan_type: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
):
    sql, extra_params = _build_loans_query(
        search=search,
        status=status,
        loan_type=loan_type,
        date_from=date_from,
        date_to=date_to,
    )

    rows = db.execute(
        sql,
        {"tenant_id": membership.tenant_id, **extra_params},
    ).mappings().all()

    items = []
    today = date.today()

    for row in rows:
        actual_return_date = row["actual_return_date"]
        expected_return_date = row["expected_return_date"]

        if actual_return_date:
            effective_status = "RETURNED"
        elif expected_return_date and expected_return_date < today:
            effective_status = "LATE"
        else:
            effective_status = "ACTIVE"

        items.append(
            {
                "id": str(row["id"]),
                "start_date": row["start_date"].isoformat() if row["start_date"] else None,
                "expected_return_date": expected_return_date.isoformat() if expected_return_date else None,
                "actual_return_date": actual_return_date.isoformat() if actual_return_date else None,
                "status": row["status"],
                "effective_status": effective_status,
                "loan_type": row["loan_type"] or "LOAN",
                "amount": float(row["amount"]) if row["amount"] is not None else None,
                "customer_name": row["customer_name"].strip() if row["customer_name"] else "",
                "notes": row["notes"],
                "dress_code": row["dress_code"],
                "dress_name": row["dress_name"],
            }
        )

    return {
        "items": items,
        "total": len(items),
    }


@router.get("/loans/export")
def export_loans_report(
    db: Session = Depends(get_db),
    membership=Depends(require_roles("admin", "manager", "staff")),
    search: str | None = Query(default=None),
    status: str | None = Query(default=None),
    loan_type: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    lang: str = Query(default="es"),
):
    sql, extra_params = _build_loans_query(
        search=search,
        status=status,
        loan_type=loan_type,
        date_from=date_from,
        date_to=date_to,
    )

    rows = db.execute(
        sql,
        {"tenant_id": membership.tenant_id, **extra_params},
    ).mappings().all()

    tr = _tx("loans", lang)

    wb = Workbook()
    ws = wb.active
    ws.title = tr["sheet"]

    _apply_header_style(ws, tr["headers"])

    today = date.today()

    for row in rows:
        customer_name = row["customer_name"].strip() if row["customer_name"] else ""
        actual_return_date = row["actual_return_date"]
        expected_return_date = row["expected_return_date"]

        if actual_return_date:
            effective_status = "RETURNED"
        elif expected_return_date and expected_return_date < today:
            effective_status = "LATE"
        else:
            effective_status = "ACTIVE"

        ws.append(
            [
                _loan_type_label(row["loan_type"], lang),
                row["start_date"].strftime("%Y-%m-%d") if row["start_date"] else "",
                expected_return_date.strftime("%Y-%m-%d") if expected_return_date else "",
                actual_return_date.strftime("%Y-%m-%d") if actual_return_date else "",
                _loan_status_label(effective_status, lang),
                row["dress_code"],
                row["dress_name"],
                customer_name,
                float(row["amount"]) if row["amount"] is not None else None,
                row["notes"],
            ]
        )

        status_cell = ws.cell(row=ws.max_row, column=5)
        status_cell.fill = _loan_status_fill(effective_status)
        status_cell.font = _loan_status_font(effective_status)
        status_cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=9, max_col=9):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 24
    ws.column_dimensions["H"].width = 28
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 32

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{ws.max_row}"

    return _excel_response(wb, tr["filename"])


# =========================================================
# COSTOS DE PRODUCCION# =========================================================
# COSTOS DE PRODUCCION
# =========================================================

def _build_production_costs_query(
    search: str | None = None,
    status: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
):
    sql = """
        SELECT
            po.id,
            po.order_number,
            COALESCE(po.target_dress_name, '-') AS target_dress_name,
            COALESCE(po.target_dress_code, '-') AS target_dress_code,
            COALESCE(s.name, '-') AS workshop_supplier_name,
            po.status,
            po.priority,
            po.due_date,
            po.planned_quantity,
            po.produced_quantity,
            COALESCE(po.labor_cost, 0) AS labor_cost,
            COALESCE(po.additional_cost, 0) AS additional_cost,
            COALESCE(po.currency, 'USD') AS currency,

            COALESCE(
                SUM(
                    COALESCE(m.planned_quantity, 0) * COALESCE(m.unit_cost_snapshot, 0)
                ),
                0
            ) AS estimated_material_cost,

            COALESCE(
                SUM(
                    (
                        COALESCE(m.delivered_quantity, 0)
                        - COALESCE(m.returned_quantity, 0)
                        - COALESCE(m.waste_quantity, 0)
                    ) * COALESCE(m.unit_cost_snapshot, 0)
                ),
                0
            ) AS actual_material_cost

        FROM production_orders po
        LEFT JOIN suppliers s
            ON s.id = po.workshop_supplier_id
        LEFT JOIN production_order_materials m
            ON m.production_order_id = po.id

        WHERE po.tenant_id = :tenant_id
          AND po.deleted_at IS NULL
    """

    params: dict[str, object] = {}

    if status:
        sql += " AND po.status = :status"
        params["status"] = status.strip().upper()

    if date_from:
        sql += " AND DATE(po.due_date) >= :date_from"
        params["date_from"] = date_from

    if date_to:
        sql += " AND DATE(po.due_date) <= :date_to"
        params["date_to"] = date_to

    if search:
        sql += """
          AND (
            LOWER(COALESCE(po.order_number, '')) LIKE :search
            OR LOWER(COALESCE(po.target_dress_name, '')) LIKE :search
            OR LOWER(COALESCE(po.target_dress_code, '')) LIKE :search
          )
        """
        params["search"] = f"%{search.strip().lower()}%"

    sql += """
        GROUP BY
            po.id,
            po.order_number,
            po.target_dress_name,
            po.target_dress_code,
            s.name,
            po.status,
            po.priority,
            po.due_date,
            po.planned_quantity,
            po.produced_quantity,
            po.labor_cost,
            po.additional_cost,
            po.currency
        ORDER BY po.due_date DESC NULLS LAST, po.order_number DESC
    """

    return text(sql), params


@router.get("/production-costs")
def production_costs_report(
    db: Session = Depends(get_db),
    membership=Depends(require_roles("admin", "manager", "staff")),
    search: str | None = Query(default=None),
    status: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
):
    sql, extra_params = _build_production_costs_query(
        search=search,
        status=status,
        date_from=date_from,
        date_to=date_to,
    )

    rows = db.execute(
        sql,
        {"tenant_id": membership.tenant_id, **extra_params},
    ).mappings().all()

    items = []

    for row in rows:
        estimated_material = Decimal(str(row["estimated_material_cost"] or 0))
        actual_material = Decimal(str(row["actual_material_cost"] or 0))
        labor = Decimal(str(row["labor_cost"] or 0))
        additional = Decimal(str(row["additional_cost"] or 0))

        total_estimated = estimated_material + labor + additional
        total_actual = actual_material + labor + additional

        planned = Decimal(str(row["planned_quantity"] or 0))
        produced = Decimal(str(row["produced_quantity"] or 0))

        unit_estimated = float(total_estimated / planned) if planned > 0 else None
        unit_actual = float(total_actual / produced) if produced > 0 else None

        items.append(
            {
                "id": str(row["id"]),
                "order_number": row["order_number"],
                "dress": row["target_dress_name"],
                "workshop": row["workshop_supplier_name"],
                "status": row["status"],
                "priority": row["priority"],
                "due_date": row["due_date"].isoformat() if row["due_date"] else None,
                "planned_quantity": float(planned),
                "produced_quantity": float(produced),
                "estimated_material_cost": float(estimated_material),
                "actual_material_cost": float(actual_material),
                "labor_cost": float(labor),
                "additional_cost": float(additional),
                "total_estimated": float(total_estimated),
                "total_actual": float(total_actual),
                "unit_estimated": unit_estimated,
                "unit_actual": unit_actual,
                "currency": "ARS",
            }
        )

    return {
        "items": items,
        "total": len(items),
    }


@router.get("/production-costs/export")
def export_production_costs_report(
    db: Session = Depends(get_db),
    membership=Depends(require_roles("admin", "manager", "staff")),
    search: str | None = Query(default=None),
    status: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    lang: str = Query(default="es"),
):
    sql, extra_params = _build_production_costs_query(
        search=search,
        status=status,
        date_from=date_from,
        date_to=date_to,
    )

    rows = db.execute(
        sql,
        {"tenant_id": membership.tenant_id, **extra_params},
    ).mappings().all()

    tr = _tx("production_costs", lang)

    wb = Workbook()
    ws = wb.active
    ws.title = tr["sheet"]

    _apply_header_style(ws, tr["headers"])

    for row in rows:
        estimated_material = Decimal(str(row["estimated_material_cost"] or 0))
        actual_material = Decimal(str(row["actual_material_cost"] or 0))
        labor = Decimal(str(row["labor_cost"] or 0))
        additional = Decimal(str(row["additional_cost"] or 0))

        total_estimated = estimated_material + labor + additional
        total_actual = actual_material + labor + additional

        planned = Decimal(str(row["planned_quantity"] or 0))
        produced = Decimal(str(row["produced_quantity"] or 0))

        unit_estimated = float(total_estimated / planned) if planned > 0 else None
        unit_actual = float(total_actual / produced) if produced > 0 else None

        raw_status = (row["status"] or "").upper()
        raw_priority = (row["priority"] or "").upper()

        ws.append(
            [
                row["order_number"],
                row["target_dress_name"],
                row["workshop_supplier_name"],
                tr["status"].get(raw_status, row["status"]),
                tr["priority"].get(raw_priority, row["priority"]),
                row["due_date"].strftime("%Y-%m-%d") if row["due_date"] else "",
                float(planned),
                float(produced),
                float(estimated_material),
                float(actual_material),
                float(labor),
                float(additional),
                float(total_estimated),
                float(total_actual),
                unit_estimated,
                unit_actual,
                "ARS",
            ]
        )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=16):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 16
    ws.column_dimensions["K"].width = 16
    ws.column_dimensions["L"].width = 14
    ws.column_dimensions["M"].width = 16
    ws.column_dimensions["N"].width = 16
    ws.column_dimensions["O"].width = 16
    ws.column_dimensions["P"].width = 16
    ws.column_dimensions["Q"].width = 12

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:Q{ws.max_row}"

    return _excel_response(wb, tr["filename"])


# =========================================================
# VENTAS DE VESTIDOS# =========================================================
# VENTAS DE VESTIDOS
# =========================================================

@router.get("/sales")
def sales_report(
    q: str | None = Query(default=None),
    payment_method: str | None = Query(default=None),
    status_filter: str | None = Query(default=None, alias="status"),
    date_from: datetime | None = Query(default=None),
    date_to: datetime | None = Query(default=None),
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
):
    tenant_id = membership.tenant_id

    stmt = (
        select(DressSale, Dress, Customer)
        .join(Dress, Dress.id == DressSale.dress_id)
        .outerjoin(Customer, Customer.id == DressSale.customer_id)
        .where(DressSale.tenant_id == tenant_id)
    )

    if status_filter:
        stmt = stmt.where(DressSale.status == status_filter)
    if payment_method:
        stmt = stmt.where(DressSale.payment_method == payment_method)
    if date_from:
        stmt = stmt.where(DressSale.sale_date >= date_from)
    if date_to:
        stmt = stmt.where(DressSale.sale_date <= date_to)
    if q:
        pattern = f"%{q.strip()}%"
        stmt = stmt.where(
            or_(
                Dress.code.ilike(pattern),
                Dress.name.ilike(pattern),
                Customer.first_name.ilike(pattern),
                Customer.last_name.ilike(pattern),
            )
        )

    rows = db.execute(stmt.order_by(desc(DressSale.sale_date))).all()

    items = []
    total_amount = 0

    for sale, dress, customer in rows:
        total_amount += float(sale.sale_price or 0)
        items.append({
            "id": str(sale.id),
            "sale_number": getattr(sale, "sale_number", None),
            "sale_date": sale.sale_date.isoformat() if sale.sale_date else None,
            "dress_code": getattr(dress, "code", None),
            "dress_name": getattr(dress, "name", None),
            "customer_full_name": getattr(customer, "full_name", None) or (
                f"{getattr(customer, 'first_name', '')} {getattr(customer, 'last_name', '')}".strip() if customer else ""
            ),
            "sale_price": float(sale.sale_price or 0),
            "currency": sale.currency,
            "payment_method": _payment_method_label(sale.payment_method),
            "status": sale.status,
            "notes": sale.notes,
        })

    return {
        "items": items,
        "total": len(items),
        "total_amount": total_amount,
    }

@router.get("/sales/export")
def export_sales_report(
    q: str | None = Query(default=None),
    payment_method: str | None = Query(default=None),
    status_filter: str | None = Query(default=None, alias="status"),
    date_from: datetime | None = Query(default=None),
    date_to: datetime | None = Query(default=None),
    lang: str = Query(default="es"),
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
):
    tenant_id = membership.tenant_id

    stmt = (
        select(DressSale, Dress, Customer)
        .join(Dress, Dress.id == DressSale.dress_id)
        .outerjoin(Customer, Customer.id == DressSale.customer_id)
        .where(DressSale.tenant_id == tenant_id)
    )

    if status_filter:
        stmt = stmt.where(DressSale.status == status_filter)
    if payment_method:
        stmt = stmt.where(DressSale.payment_method == payment_method)
    if date_from:
        stmt = stmt.where(DressSale.sale_date >= date_from)
    if date_to:
        stmt = stmt.where(DressSale.sale_date <= date_to)
    if q:
        pattern = f"%{q.strip()}%"
        stmt = stmt.where(
            or_(
                Dress.code.ilike(pattern),
                Dress.name.ilike(pattern),
                Customer.first_name.ilike(pattern),
                Customer.last_name.ilike(pattern),
            )
        )

    rows = db.execute(stmt.order_by(desc(DressSale.sale_date))).all()

    tr = _tx("sales", lang)

    wb = Workbook()
    ws = wb.active
    ws.title = tr["sheet"]

    _apply_header_style(ws, tr["headers"])

    total_amount = Decimal("0")

    for sale, dress, customer in rows:
        sale_price = Decimal(str(sale.sale_price or 0))
        total_amount += sale_price

        customer_name = getattr(customer, "full_name", None) or (
            f"{getattr(customer, 'first_name', '')} {getattr(customer, 'last_name', '')}".strip()
            if customer else ""
        )

        ws.append(
            [
                getattr(sale, "sale_number", "") or "",
                sale.sale_date.strftime("%Y-%m-%d %H:%M") if sale.sale_date else "",
                getattr(dress, "code", "") or "",
                getattr(dress, "name", "") or "",
                customer_name,
                float(sale_price),
                sale.currency or "USD",
                _payment_method_label(sale.payment_method, lang),
                _sale_status_label(sale.status, lang),
                sale.notes or "",
            ]
        )

        current_row = ws.max_row

        price_cell = ws.cell(row=current_row, column=6)
        price_cell.number_format = "#,##0.00"
        price_cell.alignment = Alignment(horizontal="right", vertical="center")

        currency_cell = ws.cell(row=current_row, column=7)
        currency_cell.alignment = Alignment(horizontal="center", vertical="center")

        payment_cell = ws.cell(row=current_row, column=8)
        payment_cell.alignment = Alignment(horizontal="center", vertical="center")

        status_cell = ws.cell(row=current_row, column=9)
        status_cell.fill = _sale_status_fill(sale.status)
        status_cell.font = _sale_status_font(sale.status)
        status_cell.alignment = Alignment(horizontal="center", vertical="center")

        notes_cell = ws.cell(row=current_row, column=10)
        notes_cell.alignment = Alignment(wrap_text=True, vertical="top")

    total_row_idx = ws.max_row + 2
    ws.cell(row=total_row_idx, column=8, value=tr["total"])
    ws.cell(row=total_row_idx, column=8).font = Font(bold=True)
    ws.cell(row=total_row_idx, column=8).alignment = Alignment(horizontal="right", vertical="center")

    total_value_cell = ws.cell(row=total_row_idx, column=9, value=float(total_amount))
    total_value_cell.font = Font(bold=True)
    total_value_cell.number_format = "#,##0.00"
    total_value_cell.alignment = Alignment(horizontal="right", vertical="center")

    total_label_fill = PatternFill("solid", fgColor="F4F1F5")
    ws.cell(row=total_row_idx, column=8).fill = total_label_fill
    ws.cell(row=total_row_idx, column=9).fill = total_label_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{ws.max_row}"

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 22
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 40

    return _excel_response(wb, tr["filename"])

# =========================================================
# VENTAS DE ACCESORIOS# =========================================================
# VENTAS DE ACCESORIOS
# =========================================================

@router.get("/accessory-sales")
def accessory_sales_report(
    q: str | None = Query(default=None),
    payment_method: str | None = Query(default=None),
    status_filter: str | None = Query(default=None, alias="status"),
    date_from: datetime | None = Query(default=None),
    date_to: datetime | None = Query(default=None),
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
):
    tenant_id = membership.tenant_id

    stmt = (
        select(AccessorySale, Accessory, Customer)
        .join(Accessory, Accessory.id == AccessorySale.accessory_id)
        .outerjoin(Customer, Customer.id == AccessorySale.customer_id)
        .where(AccessorySale.tenant_id == tenant_id)
    )

    if status_filter:
        stmt = stmt.where(AccessorySale.status == status_filter)

    if payment_method:
        stmt = stmt.where(AccessorySale.payment_method == payment_method)

    if date_from:
        stmt = stmt.where(AccessorySale.sale_date >= date_from)

    if date_to:
        stmt = stmt.where(AccessorySale.sale_date <= date_to)

    if q:
        pattern = f"%{q.strip()}%"
        search_clauses = [
            Accessory.code.ilike(pattern),
            Accessory.name.ilike(pattern),
            Customer.first_name.ilike(pattern),
            Customer.last_name.ilike(pattern),
        ]

        if hasattr(AccessorySale, "sale_number"):
            search_clauses.append(AccessorySale.sale_number.ilike(pattern))

        stmt = stmt.where(or_(*search_clauses))

    rows = db.execute(stmt.order_by(desc(AccessorySale.sale_date))).all()

    items = []
    total_amount = Decimal("0")

    for sale, accessory, customer in rows:
        total_amount += Decimal(str(sale.total_price or 0))

        items.append(
            {
                "id": str(sale.id),
                "sale_number": getattr(sale, "sale_number", None),
                "sale_date": sale.sale_date.isoformat() if sale.sale_date else None,
                "accessory_code": getattr(accessory, "code", None),
                "accessory_name": getattr(accessory, "name", None),
                "customer_full_name": getattr(customer, "full_name", None) or (
                    f"{getattr(customer, 'first_name', '')} {getattr(customer, 'last_name', '')}".strip()
                    if customer else ""
                ),
                "quantity": int(sale.quantity or 0),
                "unit_price": float(sale.unit_price or 0),
                "total_price": float(sale.total_price or 0),
                "currency": sale.currency or "ARS",
                "payment_method": _payment_method_label(sale.payment_method),
                "status": sale.status,
                "notes": sale.notes,
            }
        )

    return {
        "items": items,
        "total": len(items),
        "total_amount": float(total_amount),
    }

@router.get("/accessory-sales/export")
def export_accessory_sales_report(
    q: str | None = Query(default=None),
    payment_method: str | None = Query(default=None),
    status_filter: str | None = Query(default=None, alias="status"),
    date_from: datetime | None = Query(default=None),
    date_to: datetime | None = Query(default=None),
    lang: str = Query(default="es"),
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
):
    tenant_id = membership.tenant_id

    stmt = (
        select(AccessorySale, Accessory, Customer)
        .join(Accessory, Accessory.id == AccessorySale.accessory_id)
        .outerjoin(Customer, Customer.id == AccessorySale.customer_id)
        .where(AccessorySale.tenant_id == tenant_id)
    )

    if status_filter:
        stmt = stmt.where(AccessorySale.status == status_filter)

    if payment_method:
        stmt = stmt.where(AccessorySale.payment_method == payment_method)

    if date_from:
        stmt = stmt.where(AccessorySale.sale_date >= date_from)

    if date_to:
        stmt = stmt.where(AccessorySale.sale_date <= date_to)

    if q:
        pattern = f"%{q.strip()}%"
        search_clauses = [
            Accessory.code.ilike(pattern),
            Accessory.name.ilike(pattern),
            Customer.first_name.ilike(pattern),
            Customer.last_name.ilike(pattern),
        ]

        if hasattr(AccessorySale, "sale_number"):
            search_clauses.append(AccessorySale.sale_number.ilike(pattern))

        stmt = stmt.where(or_(*search_clauses))

    rows = db.execute(stmt.order_by(desc(AccessorySale.sale_date))).all()

    tr = _tx("accessory_sales", lang)

    wb = Workbook()
    ws = wb.active
    ws.title = tr["sheet"]

    _apply_header_style(ws, tr["headers"])

    total_amount = Decimal("0")

    for sale, accessory, customer in rows:
        total_price = Decimal(str(sale.total_price or 0))
        total_amount += total_price

        customer_name = getattr(customer, "full_name", None) or (
            f"{getattr(customer, 'first_name', '')} {getattr(customer, 'last_name', '')}".strip()
            if customer else ""
        )

        ws.append(
            [
                getattr(sale, "sale_number", "") or "",
                sale.sale_date.strftime("%Y-%m-%d %H:%M") if sale.sale_date else "",
                getattr(accessory, "code", "") or "",
                getattr(accessory, "name", "") or "",
                customer_name,
                int(sale.quantity or 0),
                float(sale.unit_price or 0),
                float(total_price),
                sale.currency or "ARS",
                _payment_method_label(sale.payment_method, lang),
                _sale_status_label(sale.status, lang),
                sale.notes or "",
            ]
        )

        current_row = ws.max_row

        qty_cell = ws.cell(row=current_row, column=6)
        qty_cell.alignment = Alignment(horizontal="center", vertical="center")

        unit_price_cell = ws.cell(row=current_row, column=7)
        unit_price_cell.number_format = "#,##0.00"
        unit_price_cell.alignment = Alignment(horizontal="right", vertical="center")

        total_cell = ws.cell(row=current_row, column=8)
        total_cell.number_format = "#,##0.00"
        total_cell.alignment = Alignment(horizontal="right", vertical="center")

        currency_cell = ws.cell(row=current_row, column=9)
        currency_cell.alignment = Alignment(horizontal="center", vertical="center")

        payment_cell = ws.cell(row=current_row, column=10)
        payment_cell.alignment = Alignment(horizontal="center", vertical="center")

        status_cell = ws.cell(row=current_row, column=11)
        status_cell.fill = _sale_status_fill(sale.status)
        status_cell.font = _sale_status_font(sale.status)
        status_cell.alignment = Alignment(horizontal="center", vertical="center")

        notes_cell = ws.cell(row=current_row, column=12)
        notes_cell.alignment = Alignment(wrap_text=True, vertical="top")

    total_row_idx = ws.max_row + 2

    ws.cell(row=total_row_idx, column=10, value=tr["total"])
    ws.cell(row=total_row_idx, column=10).font = Font(bold=True)
    ws.cell(row=total_row_idx, column=10).alignment = Alignment(horizontal="right", vertical="center")

    total_value_cell = ws.cell(row=total_row_idx, column=11, value=float(total_amount))
    total_value_cell.font = Font(bold=True)
    total_value_cell.number_format = "#,##0.00"
    total_value_cell.alignment = Alignment(horizontal="right", vertical="center")

    total_label_fill = PatternFill("solid", fgColor="F4F1F5")
    ws.cell(row=total_row_idx, column=10).fill = total_label_fill
    ws.cell(row=total_row_idx, column=11).fill = total_label_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{ws.max_row}"

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 22
    ws.column_dimensions["K"].width = 14
    ws.column_dimensions["L"].width = 40

    return _excel_response(wb, tr["filename"])


# =========================================================
# VENTAS UNIFICADAS# =========================================================
# VENTAS UNIFICADAS
# =========================================================

def _build_sales_unified_query(
    search: str | None = None,
    status: str | None = None,
    currency: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
):
    sql = """
        SELECT
            s.id,
            s.sale_number,
            s.sale_date,
            s.customer_id,
            s.currency,
            s.status,
            s.subtotal_amount,
            s.discount_amount,
            s.total_amount,
            s.notes,
            TRIM(COALESCE(c.first_name, '') || ' ' || COALESCE(c.last_name, '')) AS customer_full_name
        FROM sales s
        LEFT JOIN customers c
            ON c.id = s.customer_id
        WHERE s.tenant_id = :tenant_id
    """

    params: dict[str, object] = {}

    if status:
        sql += " AND s.status = :status"
        params["status"] = status.strip().upper()

    if currency:
        sql += """
          AND EXISTS (
            SELECT 1
            FROM sale_items si_currency
            WHERE si_currency.sale_id = s.id
              AND si_currency.tenant_id = s.tenant_id
              AND UPPER(COALESCE(si_currency.currency, 'ARS')) = :currency
          )
        """
        params["currency"] = currency.strip().upper()

    if date_from:
        sql += " AND DATE(s.sale_date) >= :date_from"
        params["date_from"] = date_from

    if date_to:
        sql += " AND DATE(s.sale_date) <= :date_to"
        params["date_to"] = date_to

    if search:
        sql += """
          AND (
            LOWER(COALESCE(s.sale_number, '')) LIKE :search
            OR LOWER(COALESCE(c.first_name, '')) LIKE :search
            OR LOWER(COALESCE(c.last_name, '')) LIKE :search
            OR LOWER(COALESCE(s.notes, '')) LIKE :search
          )
        """
        params["search"] = f"%{search.strip().lower()}%"

    sql += """
        ORDER BY s.sale_date DESC, s.created_at DESC
    """

    return text(sql), params


def _get_sales_unified_items(db: Session, tenant_id, sale_id):
    sql = text("""
        SELECT
            si.id,
            si.item_type,
            si.description_snapshot,
            si.quantity,
            si.unit_price,
            si.currency,
            si.line_total
        FROM sale_items si
        WHERE si.tenant_id = :tenant_id
          AND si.sale_id = :sale_id
        ORDER BY si.created_at ASC, si.id ASC
    """)
    return db.execute(sql, {"tenant_id": tenant_id, "sale_id": sale_id}).mappings().all()


def _get_sales_unified_payments(db: Session, tenant_id, sale_id):
    sql = text("""
        SELECT
            sp.id,
            sp.payment_method,
            sp.amount,
            sp.currency,
            sp.reference,
            sp.notes
        FROM sale_payments sp
        WHERE sp.tenant_id = :tenant_id
          AND sp.sale_id = :sale_id
        ORDER BY sp.created_at ASC, sp.id ASC
    """)
    return db.execute(sql, {"tenant_id": tenant_id, "sale_id": sale_id}).mappings().all()

@router.get("/dress-stock-valuation")
def dress_stock_valuation_report(
    db: Session = Depends(get_db),
    membership=Depends(require_roles("admin", "manager", "staff")),
    search: str | None = Query(default=None),
):
    sql = """
        SELECT
            d.id,
            d.code,
            d.name,
            d.size,
            d.color,
            d.status,
            d.sale_price,
            d.rental_price,
            c.name AS capsule_name
        FROM dresses d
        LEFT JOIN capsules c ON c.id = d.capsule_id
        WHERE d.tenant_id = :tenant_id
          AND d.deleted_at IS NULL
    """

    params = {"tenant_id": membership.tenant_id}

    if search:
        sql += """
          AND (
            LOWER(d.code) LIKE :search
            OR LOWER(d.name) LIKE :search
            OR LOWER(COALESCE(d.color,'')) LIKE :search
          )
        """
        params["search"] = f"%{search.lower()}%"

    rows = db.execute(text(sql), params).mappings().all()

    items = []
    total_sale = Decimal("0")
    total_rental = Decimal("0")
    total_items = 0
    available_items = 0

    for r in rows:
        if r["status"] != "SOLD":
            total_items += 1

            sale_price = Decimal(str(r["sale_price"] or 0))
            rental_price = Decimal(str(r["rental_price"] or 0))

            total_sale += sale_price
            total_rental += rental_price

            if r["status"] == "AVAILABLE":
                available_items += 1

            items.append({
                "id": str(r["id"]),
                "code": r["code"],
                "name": r["name"],
                "capsule": r["capsule_name"],
                "size": r["size"],
                "color": r["color"],
                "status": r["status"],
                "sale_price": float(sale_price),
                "rental_price": float(rental_price),
            })

    return {
        "items": items,
        "total": len(items),
        "kpis": {
            "total_items": total_items,
            "available_items": available_items,
            "total_sale_value": float(total_sale),
            "total_rental_value": float(total_rental),
        }
    }

def _dress_status_label(value: str | None) -> str:
    raw = (value or "").upper()
    if raw == "AVAILABLE":
        return "Disponible"
    if raw == "LOANED":
        return "Prestado"
    if raw == "RENTED":
        return "Alquilado"
    if raw == "CLEANING":
        return "Limpieza"
    if raw == "MAINTENANCE":
        return "Reparación"
    if raw == "SOLD":
        return "Vendido"
    if raw == "RETIRED":
        return "Retirado"
    return value or ""


@router.get("/dress-stock-valuation/export")
def export_dress_stock_valuation_report(
    db: Session = Depends(get_db),
    membership=Depends(require_roles("admin", "manager", "staff")),
    search: str | None = Query(default=None),
):
    sql = """
        SELECT
            d.code,
            d.name,
            c.name AS capsule,
            d.size,
            d.color,
            d.status,
            COALESCE(d.sale_price, 0) AS sale_price,
            COALESCE(d.rental_price, 0) AS rental_price
        FROM dresses d
        LEFT JOIN capsules c
            ON c.id = d.capsule_id
        WHERE d.tenant_id = :tenant_id
          AND d.deleted_at IS NULL
          AND d.status != 'SOLD'
    """

    params: dict[str, object] = {
        "tenant_id": membership.tenant_id,
    }

    if search:
        sql += """
          AND (
            LOWER(COALESCE(d.code, '')) LIKE :search
            OR LOWER(COALESCE(d.name, '')) LIKE :search
            OR LOWER(COALESCE(c.name, '')) LIKE :search
            OR LOWER(COALESCE(d.size, '')) LIKE :search
            OR LOWER(COALESCE(d.color, '')) LIKE :search
          )
        """
        params["search"] = f"%{search.strip().lower()}%"

    sql += """
        ORDER BY d.name ASC, d.code ASC
    """

    rows = db.execute(text(sql), params).mappings().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock vestidos"

    headers = [
        "Código",
        "Vestido",
        "Cápsula",
        "Talle",
        "Color",
        "Estado",
        "Precio venta (USD)",
        "Precio alquiler (USD)",
    ]
    _apply_header_style(ws, headers)

    total_sale = Decimal("0")
    total_rental = Decimal("0")

    for row in rows:
        sale_price = Decimal(str(row["sale_price"] or 0))
        rental_price = Decimal(str(row["rental_price"] or 0))

        total_sale += sale_price
        total_rental += rental_price

        ws.append(
            [
                row["code"],
                row["name"],
                row["capsule"],
                row["size"],
                row["color"],
                _dress_status_label(row["status"]),
                float(sale_price),
                float(rental_price),
            ]
        )

    total_row_idx = ws.max_row + 2

    ws.cell(row=total_row_idx, column=6, value="TOTAL")
    ws.cell(row=total_row_idx, column=6).font = Font(bold=True)
    ws.cell(row=total_row_idx, column=6).alignment = Alignment(
        horizontal="right",
        vertical="center",
    )

    ws.cell(row=total_row_idx, column=7, value=float(total_sale))
    ws.cell(row=total_row_idx, column=7).font = Font(bold=True)

    ws.cell(row=total_row_idx, column=8, value=float(total_rental))
    ws.cell(row=total_row_idx, column=8).font = Font(bold=True)

    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=8):
        for cell in row_cells:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00 "USD"'
                cell.alignment = Alignment(horizontal="right", vertical="center")

    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
        for cell in row_cells:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 22

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"

    return _excel_response(wb, "stock_valorizado_vestidos.xlsx")

@router.get("/sales-unified")
def sales_unified_report(
    db: Session = Depends(get_db),
    membership=Depends(require_roles("admin", "manager", "staff")),
    q: str | None = Query(default=None),
    status: str | None = Query(default=None),
    currency: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
):
    sql, extra_params = _build_sales_unified_query(
        search=q,
        status=status,
        currency=currency,
        date_from=date_from,
        date_to=date_to,
    )

    sales_rows = db.execute(
        sql,
        {"tenant_id": membership.tenant_id, **extra_params},
    ).mappings().all()

    items: list[dict] = []
    currency_totals: dict[str, float] = {}
    payment_totals: dict[str, float] = {}
    mixed_count = 0

    for row in sales_rows:
        sale_id = row["id"]

        sale_items = _get_sales_unified_items(db, membership.tenant_id, sale_id)
        sale_payments = _get_sales_unified_payments(db, membership.tenant_id, sale_id)

        items_totals: dict[str, float] = {}
        paid_totals: dict[str, float] = {}

        for item in sale_items:
            currency_value = str(item["currency"] or "ARS").upper()
            line_total = float(item["line_total"] or 0)
            items_totals[currency_value] = round(
                items_totals.get(currency_value, 0) + line_total,
                2,
            )
            currency_totals[currency_value] = round(
                currency_totals.get(currency_value, 0) + line_total,
                2,
            )

        for payment in sale_payments:
            currency_value = str(payment["currency"] or "ARS").upper()
            amount = float(payment["amount"] or 0)
            paid_totals[currency_value] = round(
                paid_totals.get(currency_value, 0) + amount,
                2,
            )
            payment_totals[currency_value] = round(
                payment_totals.get(currency_value, 0) + amount,
                2,
            )

        if len([value for value in items_totals.values() if NumberLike(value) > 0]) > 1:
            mixed_count += 1

        items.append(
            {
                "id": str(row["id"]),
                "sale_number": row["sale_number"],
                "sale_date": row["sale_date"].isoformat() if row["sale_date"] else None,
                "customer_full_name": row["customer_full_name"] or None,
                "currency": row["currency"],
                "status": row["status"],
                "subtotal_amount": float(row["subtotal_amount"] or 0),
                "discount_amount": float(row["discount_amount"] or 0),
                "total_amount": float(row["total_amount"] or 0),
                "notes": row["notes"],
                "items_total_ars": float(items_totals.get("ARS", 0)),
                "items_total_usd": float(items_totals.get("USD", 0)),
                "paid_total_ars": float(paid_totals.get("ARS", 0)),
                "paid_total_usd": float(paid_totals.get("USD", 0)),
                "items_totals": items_totals,
                "paid_totals": paid_totals,
                "items": [
                    {
                        "id": str(item["id"]),
                        "item_type": item["item_type"],
                        "description_snapshot": item["description_snapshot"],
                        "quantity": int(item["quantity"] or 0),
                        "unit_price": float(item["unit_price"] or 0),
                        "currency": str(item["currency"] or "ARS").upper(),
                        "line_total": float(item["line_total"] or 0),
                    }
                    for item in sale_items
                ],
                "payments": [
                    {
                        "id": str(payment["id"]),
                        "payment_method": payment["payment_method"],
                        "amount": float(payment["amount"] or 0),
                        "currency": str(payment["currency"] or "ARS").upper(),
                        "reference": payment["reference"],
                        "notes": payment["notes"],
                    }
                    for payment in sale_payments
                ],
            }
        )

    return {
        "items": items,
        "total": len(items),
        "total_ars": float(currency_totals.get("ARS", 0)),
        "total_usd": float(currency_totals.get("USD", 0)),
        "mixed_count": mixed_count,
        "currency_totals": currency_totals,
        "payment_totals": payment_totals,
    }


def _sales_currency_order(currencies: set[str]) -> list[str]:
    priority = ["ARS", "USD", "EUR"]
    ordered = [currency for currency in priority if currency in currencies]
    ordered.extend(sorted(currency for currency in currencies if currency not in priority))
    return ordered


def _sales_money_label(currency: str, amount: Decimal | float | int) -> str:
    return f"{currency} {float(amount or 0):,.2f}"


def NumberLike(value) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


@router.get("/sales-unified/export")
def export_sales_unified_report(
    db: Session = Depends(get_db),
    membership=Depends(require_roles("admin", "manager", "staff")),
    q: str | None = Query(default=None),
    status: str | None = Query(default=None),
    currency: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    lang: str = Query(default="es"),
):
    sql, extra_params = _build_sales_unified_query(
        search=q,
        status=status,
        currency=currency,
        date_from=date_from,
        date_to=date_to,
    )

    rows = db.execute(
        sql,
        {"tenant_id": membership.tenant_id, **extra_params},
    ).mappings().all()

    tr = _tx("sales_unified", lang)

    prepared_rows: list[dict] = []
    currencies: set[str] = set()
    total_items_by_currency: dict[str, Decimal] = {}
    total_paid_by_currency: dict[str, Decimal] = {}

    for row in rows:
        sale_id = row["id"]
        item_rows = _get_sales_unified_items(db, membership.tenant_id, sale_id)
        payment_rows = _get_sales_unified_payments(db, membership.tenant_id, sale_id)

        items_totals: dict[str, Decimal] = {}
        paid_totals: dict[str, Decimal] = {}

        item_parts: list[str] = []
        for item in item_rows:
            line_total = Decimal(str(item["line_total"] or 0))
            currency_value = str(item["currency"] or "ARS").upper()
            currencies.add(currency_value)

            items_totals[currency_value] = items_totals.get(currency_value, Decimal("0")) + line_total
            total_items_by_currency[currency_value] = total_items_by_currency.get(currency_value, Decimal("0")) + line_total

            raw_item_type = str(item["item_type"] or "").upper()
            item_type_label = tr["item_types"].get(raw_item_type, tr["item_fallback"])
            item_parts.append(
                f"{item_type_label}: {item['description_snapshot'] or tr['item_fallback']} · "
                f"x{int(item['quantity'] or 0)} · {_sales_money_label(currency_value, line_total)}"
            )

        payment_parts: list[str] = []
        for payment in payment_rows:
            amount = Decimal(str(payment["amount"] or 0))
            currency_value = str(payment["currency"] or "ARS").upper()
            currencies.add(currency_value)

            paid_totals[currency_value] = paid_totals.get(currency_value, Decimal("0")) + amount
            total_paid_by_currency[currency_value] = total_paid_by_currency.get(currency_value, Decimal("0")) + amount

            reference_text = f" · {payment['reference']}" if payment["reference"] else ""
            payment_parts.append(
                f"{_payment_method_label(payment['payment_method'], lang)}: "
                f"{_sales_money_label(currency_value, amount)}{reference_text}"
            )

        prepared_rows.append(
            {
                "row": row,
                "items_totals": items_totals,
                "paid_totals": paid_totals,
                "item_parts": item_parts,
                "payment_parts": payment_parts,
            }
        )

    ordered_currencies = _sales_currency_order(currencies or {"ARS"})

    base_headers = tr["headers"][:5]
    tail_headers = tr["headers"][-3:]
    dynamic_headers = (
        base_headers
        + [f"Total ítems {currency}" if _lang(lang) == "es" else f"Items total {currency}" for currency in ordered_currencies]
        + [f"Pagado {currency}" if _lang(lang) == "es" else f"Paid {currency}" for currency in ordered_currencies]
        + tail_headers
    )

    wb = Workbook()
    ws = wb.active
    ws.title = tr["sheet"]

    _apply_header_style(ws, dynamic_headers)

    item_start_col = 6
    paid_start_col = item_start_col + len(ordered_currencies)
    text_start_col = paid_start_col + len(ordered_currencies)

    for prepared in prepared_rows:
        row = prepared["row"]
        items_totals = prepared["items_totals"]
        paid_totals = prepared["paid_totals"]

        ws.append(
            [
                row["sale_number"] or "",
                row["sale_date"].strftime("%Y-%m-%d %H:%M") if row["sale_date"] else "",
                row["customer_full_name"] or "",
                row["currency"] or "ARS",
                _sale_status_label(row["status"], lang),
                *[float(items_totals.get(currency, Decimal("0"))) for currency in ordered_currencies],
                *[float(paid_totals.get(currency, Decimal("0"))) for currency in ordered_currencies],
                " | ".join(prepared["item_parts"]),
                " | ".join(prepared["payment_parts"]),
                row["notes"] or "",
            ]
        )

        current_row = ws.max_row

        for col in range(item_start_col, text_start_col):
            cell = ws.cell(row=current_row, column=col)
            cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right", vertical="center")

        status_cell = ws.cell(row=current_row, column=5)
        status_cell.fill = _sale_status_fill(row["status"])
        status_cell.font = _sale_status_font(row["status"])
        status_cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in range(text_start_col, text_start_col + 3):
            ws.cell(row=current_row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    total_row_idx = ws.max_row + 2

    ws.cell(row=total_row_idx, column=5, value=tr["total"])
    ws.cell(row=total_row_idx, column=5).font = Font(bold=True)
    ws.cell(row=total_row_idx, column=5).alignment = Alignment(horizontal="right", vertical="center")

    for index, currency_value in enumerate(ordered_currencies):
        col = item_start_col + index
        cell = ws.cell(row=total_row_idx, column=col, value=float(total_items_by_currency.get(currency_value, Decimal("0"))))
        cell.font = Font(bold=True)
        cell.number_format = "#,##0.00"
        cell.alignment = Alignment(horizontal="right", vertical="center")

    for index, currency_value in enumerate(ordered_currencies):
        col = paid_start_col + index
        cell = ws.cell(row=total_row_idx, column=col, value=float(total_paid_by_currency.get(currency_value, Decimal("0"))))
        cell.font = Font(bold=True)
        cell.number_format = "#,##0.00"
        cell.alignment = Alignment(horizontal="right", vertical="center")

    total_fill = PatternFill("solid", fgColor="F4F1F5")
    for col in range(5, text_start_col):
        ws.cell(row=total_row_idx, column=col).fill = total_fill

    ws.freeze_panes = "A2"
    last_column_letter = ws.cell(row=1, column=text_start_col + 2).column_letter
    ws.auto_filter.ref = f"A1:{last_column_letter}{ws.max_row}"

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    for col in range(item_start_col, text_start_col):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

    ws.column_dimensions[ws.cell(row=1, column=text_start_col).column_letter].width = 52
    ws.column_dimensions[ws.cell(row=1, column=text_start_col + 1).column_letter].width = 42
    ws.column_dimensions[ws.cell(row=1, column=text_start_col + 2).column_letter].width = 30

    return _excel_response(wb, tr["filename"])

